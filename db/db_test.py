import os
import pandas as pd
import pymysql
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart.label import DataLabelList

from db.db_intent_handler import handle_db_intent

# Configuration
BASE_DIR = os.path.dirname(__file__)
STATIC_DIR = os.path.join(BASE_DIR, "static")
REPORT_PATH = os.path.join(STATIC_DIR, "track_and_trace_report.xlsx")

def generate_track_and_trace_excel():
    print("📦 Generating Track and Trace Excel Report")

    try:
        intent = "track and trace"
        message = "Track my assets"
        result = handle_db_intent(intent, message, raw=True)

        if isinstance(result, str):
            print("⚠️ Response from DB handler:", result)
            return "❌ Could not generate report. Reason: " + result

        if not isinstance(result, list) or not result:
            print("❌ No data returned for report.")
            return "❌ No track and trace data found to generate report."

        df = pd.DataFrame(result)
        os.makedirs(STATIC_DIR, exist_ok=True)

        if os.path.exists(REPORT_PATH):
            try:
                os.remove(REPORT_PATH)
            except Exception as e:
                print("⚠️ Failed to delete old file:", e)
                return f"❌ Failed to clean previous report: {e}"

        df.columns = df.columns.str.strip()
        df = df.dropna(axis=1, how='all')

        # Write Raw Data
        with pd.ExcelWriter(REPORT_PATH, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Raw Data', index=False)

        wb = load_workbook(REPORT_PATH)
        ws_raw = wb['Raw Data']

        for col in ws_raw.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws_raw.column_dimensions[col[0].column_letter].width = max_length + 2

        # Dashboard setup
        ws_dash = wb.create_sheet("Dashboard")
        ws_dash.title = "Dashboard"
        ws_dash.append(["Track & Trace Summary"])
        ws_dash.merge_cells("A1:D1")
        ws_dash["A1"].font = Font(bold=True, size=14, color="1F4E78")
        ws_dash["A1"].alignment = Alignment(horizontal="center")

        ws_dash.append(["Total Records", len(df)])
        for col in ['Location', 'Asset Type']:
            if col in df.columns:
                ws_dash.append([f"Unique {col}s", df[col].nunique()])

        for row in ws_dash.iter_rows(min_row=2, max_row=4, min_col=1, max_col=2):
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

        # Charts
        row_cursor = 6
        chart_columns = ['Asset Type', 'Location']
        chart_col = 5  # column E

        for col in chart_columns:
            match = [c for c in df.columns if c.lower() == col.lower()]
            if not match:
                continue

            col_actual = match[0]
            counts = df[col_actual].value_counts().reset_index()
            counts.columns = [col_actual, "Count"]
            if counts.empty:
                continue

            ws_dash.cell(row=row_cursor, column=1, value=f"📊 {col_actual} Distribution")
            ws_dash.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=3)
            ws_dash.cell(row=row_cursor, column=1).font = Font(bold=True, size=12, color="2F5597")
            ws_dash.cell(row=row_cursor, column=1).alignment = Alignment(horizontal="center")
            row_cursor += 1

            for r_idx, row in enumerate(dataframe_to_rows(counts, index=False, header=True), start=row_cursor):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws_dash.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == row_cursor:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

            is_skewed = counts["Count"].max() / counts["Count"].sum() >= 0.7
            chart = PieChart() if len(counts) <= 8 and not is_skewed else BarChart()
            chart.title = f"{col_actual} Chart"
            chart.height = 10
            chart.width = 14

            if isinstance(chart, BarChart):
                chart.y_axis.title = 'Count'
                chart.x_axis.title = col_actual
            else:
                chart.dataLabels = DataLabelList()
                chart.dataLabels.showPercent = True
                chart.dataLabels.showLeaderLines = True

            data = Reference(ws_dash, min_col=2, min_row=row_cursor, max_row=row_cursor + len(counts))
            labels = Reference(ws_dash, min_col=1, min_row=row_cursor + 1, max_row=row_cursor + len(counts))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            ws_dash.add_chart(chart, f"{chr(64 + chart_col)}{row_cursor}")

            row_cursor += len(counts) + 4

        for col_cells in ws_dash.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0 for cell in col_cells), default=0)
            col_letter = get_column_letter(col_cells[0].column)
            ws_dash.column_dimensions[col_letter].width = max_len + 4

        wb.save(REPORT_PATH)
        print(f"✅ Excel report saved at: {REPORT_PATH}")
        return "✅ Track and trace report generated successfully."

    except Exception as e:
        print("❌ Unexpected error during report generation:", e)
        return f"❌ Failed to generate report. Error: {e}"